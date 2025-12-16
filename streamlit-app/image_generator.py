# -*- coding: utf-8 -*-
"""
画像出力モジュール (v1.0)
ランキング表を画像として出力

方式:
1. Excelテンプレートにデータを埋め込み → 画像化（テンプレート方式）
2. matplotlibで表を描画 → PNG出力（シンプル方式）
"""

import os
import logging
from io import BytesIO
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime

import pandas as pd
import matplotlib
matplotlib.use('Agg')  # バックエンド設定（GUI不要）
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from matplotlib.table import Table
import numpy as np

# openpyxlは既にインストール済み
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from copy import copy

logger = logging.getLogger(__name__)

# 日本語フォント設定
# Windowsの場合
JAPANESE_FONTS = [
    'Yu Gothic',
    'Meiryo',
    'MS Gothic',
    'Hiragino Sans',
    'IPAGothic'
]

def setup_japanese_font():
    """日本語フォントを設定"""
    for font_name in JAPANESE_FONTS:
        try:
            # フォントが存在するかチェック
            fm.findfont(font_name, fallback_to_default=False)
            plt.rcParams['font.family'] = font_name
            logger.info(f"日本語フォント設定: {font_name}")
            return font_name
        except:
            continue

    # フォールバック
    plt.rcParams['font.family'] = 'sans-serif'
    logger.warning("日本語フォントが見つかりません")
    return None

# 初期化時にフォント設定
setup_japanese_font()

# テンプレートパス
TEMPLATE_DIR = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "_archive", "テンプレート"
)
EXCEL_TEMPLATE_PATH = os.path.join(
    TEMPLATE_DIR,
    "【テンプレ】リリース内表.xlsx"
)


class TableImageGenerator:
    """表画像生成クラス（matplotlib方式）"""

    # オリコンカラー
    COLORS = {
        'header_bg': '#003366',      # ダークブルー
        'header_text': '#FFFFFF',    # 白
        'row_bg_odd': '#FFFFFF',     # 白
        'row_bg_even': '#F0F8FF',    # 薄い青
        'rank_1_bg': '#FFD700',      # 金
        'rank_2_bg': '#C0C0C0',      # 銀
        'rank_3_bg': '#CD7F32',      # 銅
        'border': '#CCCCCC',         # グレー
        'text': '#333333',           # ダークグレー
    }

    def __init__(
        self,
        ranking_name: str,
        year: int,
        figsize: Tuple[float, float] = (8, 6),
        dpi: int = 150
    ):
        """
        Args:
            ranking_name: ランキング名
            year: 年度
            figsize: 図のサイズ（インチ）
            dpi: 解像度
        """
        self.ranking_name = ranking_name
        self.year = year
        self.figsize = figsize
        self.dpi = dpi

    def generate_overall_table_image(
        self,
        data: List[Dict],
        title: str = None,
        sample_size: int = None,
        show_rank_colors: bool = True
    ) -> BytesIO:
        """総合ランキング表の画像を生成

        Args:
            data: ランキングデータ [{rank, company, score}, ...]
            title: タイトル（省略時は自動生成）
            sample_size: 回答者数
            show_rank_colors: 順位に応じた背景色を表示

        Returns:
            BytesIO（PNG画像）
        """
        if not data:
            return None

        # データ準備
        df = pd.DataFrame(data[:10])  # TOP10
        df = df[['rank', 'company', 'score']].copy()
        df.columns = ['順位', '企業名', '得点']
        df['順位'] = df['順位'].apply(lambda x: f"{x}位" if pd.notna(x) else "-")
        df['得点'] = df['得点'].apply(lambda x: f"{x}点" if pd.notna(x) else "-")

        # 図の作成
        fig, ax = plt.subplots(figsize=self.figsize, dpi=self.dpi)
        ax.axis('off')

        # タイトル
        if title is None:
            title = f"{self.year}年 オリコン顧客満足度®調査 総合ランキング"
        subtitle = f"{self.ranking_name}"
        if sample_size:
            subtitle += f" (回答者数：{sample_size:,}名)"

        fig.suptitle(title, fontsize=14, fontweight='bold', y=0.98)
        ax.set_title(subtitle, fontsize=11, pad=10)

        # 表の作成
        table = ax.table(
            cellText=df.values,
            colLabels=df.columns,
            cellLoc='center',
            loc='center',
            colWidths=[0.15, 0.55, 0.15]
        )

        # スタイル設定
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1.2, 1.8)

        # セルの書式設定
        for (row, col), cell in table.get_celld().items():
            cell.set_edgecolor(self.COLORS['border'])

            if row == 0:
                # ヘッダー
                cell.set_facecolor(self.COLORS['header_bg'])
                cell.set_text_props(color=self.COLORS['header_text'], fontweight='bold')
            else:
                # データ行
                if show_rank_colors and col == 0:
                    # 順位列の背景色
                    rank = row
                    if rank == 1:
                        cell.set_facecolor(self.COLORS['rank_1_bg'])
                    elif rank == 2:
                        cell.set_facecolor(self.COLORS['rank_2_bg'])
                    elif rank == 3:
                        cell.set_facecolor(self.COLORS['rank_3_bg'])
                    else:
                        cell.set_facecolor(self.COLORS['row_bg_odd'] if row % 2 else self.COLORS['row_bg_even'])
                else:
                    cell.set_facecolor(self.COLORS['row_bg_odd'] if row % 2 else self.COLORS['row_bg_even'])

                cell.set_text_props(color=self.COLORS['text'])

        # 余白調整
        plt.tight_layout()
        plt.subplots_adjust(top=0.85)

        # BytesIOに保存
        output = BytesIO()
        fig.savefig(output, format='png', bbox_inches='tight', facecolor='white')
        plt.close(fig)
        output.seek(0)

        return output

    def generate_comparison_table_image(
        self,
        current_data: List[Dict],
        prev_data: List[Dict],
        title: str = None
    ) -> BytesIO:
        """前年比較付きランキング表の画像を生成

        Args:
            current_data: 今年のデータ
            prev_data: 前年のデータ
            title: タイトル

        Returns:
            BytesIO（PNG画像）
        """
        if not current_data:
            return None

        # 前年順位マッピング
        prev_ranks = {}
        if prev_data:
            for entry in prev_data:
                company = entry.get('company', '')
                prev_ranks[company] = entry.get('rank')

        # データ準備
        rows = []
        for entry in current_data[:10]:
            rank = entry.get('rank')
            company = entry.get('company', '')
            score = entry.get('score')
            prev_rank = prev_ranks.get(company)

            # 変動
            if prev_rank:
                diff = prev_rank - rank
                if diff > 0:
                    change = f"↑{diff}"
                elif diff < 0:
                    change = f"↓{abs(diff)}"
                else:
                    change = "→"
            else:
                change = "NEW"

            rows.append({
                '順位': f"{rank}位" if rank else "-",
                '前回': f"{prev_rank}位" if prev_rank else "-",
                '変動': change,
                '企業名': company,
                '得点': f"{score}点" if score else "-"
            })

        df = pd.DataFrame(rows)

        # 図の作成
        fig, ax = plt.subplots(figsize=(10, 6), dpi=self.dpi)
        ax.axis('off')

        # タイトル
        if title is None:
            title = f"{self.year}年 オリコン顧客満足度®調査 総合ランキング（前年比較）"

        fig.suptitle(title, fontsize=14, fontweight='bold', y=0.98)
        ax.set_title(self.ranking_name, fontsize=11, pad=10)

        # 表の作成
        table = ax.table(
            cellText=df.values,
            colLabels=df.columns,
            cellLoc='center',
            loc='center',
            colWidths=[0.12, 0.12, 0.1, 0.45, 0.12]
        )

        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1.2, 1.8)

        # スタイル設定
        for (row, col), cell in table.get_celld().items():
            cell.set_edgecolor(self.COLORS['border'])

            if row == 0:
                cell.set_facecolor(self.COLORS['header_bg'])
                cell.set_text_props(color=self.COLORS['header_text'], fontweight='bold')
            else:
                cell.set_facecolor(self.COLORS['row_bg_odd'] if row % 2 else self.COLORS['row_bg_even'])
                cell.set_text_props(color=self.COLORS['text'])

                # 変動列の色分け
                if col == 2:  # 変動列
                    text = cell.get_text().get_text()
                    if '↑' in text:
                        cell.set_text_props(color='green', fontweight='bold')
                    elif '↓' in text:
                        cell.set_text_props(color='red', fontweight='bold')
                    elif 'NEW' in text:
                        cell.set_text_props(color='blue', fontweight='bold')

        plt.tight_layout()
        plt.subplots_adjust(top=0.85)

        output = BytesIO()
        fig.savefig(output, format='png', bbox_inches='tight', facecolor='white')
        plt.close(fig)
        output.seek(0)

        return output

    def generate_multi_table_image(
        self,
        tables_data: Dict[str, List[Dict]],
        title: str = None
    ) -> BytesIO:
        """複数表（評価項目別/部門別）の画像を生成

        Args:
            tables_data: {表名: データリスト}
            title: タイトル

        Returns:
            BytesIO（PNG画像）
        """
        if not tables_data:
            return None

        n_tables = len(tables_data)
        if n_tables == 0:
            return None

        # レイアウト計算
        n_cols = min(3, n_tables)
        n_rows = (n_tables + n_cols - 1) // n_cols

        fig_width = 5 * n_cols
        fig_height = 4 * n_rows + 1

        fig, axes = plt.subplots(n_rows, n_cols, figsize=(fig_width, fig_height), dpi=self.dpi)

        if n_tables == 1:
            axes = [[axes]]
        elif n_rows == 1:
            axes = [axes]

        # タイトル
        if title is None:
            title = f"{self.year}年 オリコン顧客満足度®調査"
        fig.suptitle(title, fontsize=14, fontweight='bold', y=0.98)

        # 各表を描画
        table_items = list(tables_data.items())
        for idx, (name, data) in enumerate(table_items):
            row_idx = idx // n_cols
            col_idx = idx % n_cols
            ax = axes[row_idx][col_idx]
            ax.axis('off')
            ax.set_title(name, fontsize=10, fontweight='bold')

            if not data:
                continue

            # データ準備（TOP5）
            df = pd.DataFrame(data[:5])
            if 'rank' in df.columns and 'company' in df.columns and 'score' in df.columns:
                df = df[['rank', 'company', 'score']].copy()
                df.columns = ['順位', '企業名', '得点']
                df['順位'] = df['順位'].apply(lambda x: f"{x}位" if pd.notna(x) else "-")
                df['得点'] = df['得点'].apply(lambda x: f"{x}点" if pd.notna(x) else "-")

                table = ax.table(
                    cellText=df.values,
                    colLabels=df.columns,
                    cellLoc='center',
                    loc='center',
                    colWidths=[0.2, 0.55, 0.2]
                )

                table.auto_set_font_size(False)
                table.set_fontsize(8)
                table.scale(1.0, 1.5)

                for (r, c), cell in table.get_celld().items():
                    cell.set_edgecolor(self.COLORS['border'])
                    if r == 0:
                        cell.set_facecolor(self.COLORS['header_bg'])
                        cell.set_text_props(color=self.COLORS['header_text'], fontweight='bold')
                    else:
                        cell.set_facecolor(self.COLORS['row_bg_odd'] if r % 2 else self.COLORS['row_bg_even'])

        # 余った軸を非表示
        for idx in range(n_tables, n_rows * n_cols):
            row_idx = idx // n_cols
            col_idx = idx % n_cols
            axes[row_idx][col_idx].axis('off')

        plt.tight_layout()
        plt.subplots_adjust(top=0.92)

        output = BytesIO()
        fig.savefig(output, format='png', bbox_inches='tight', facecolor='white')
        plt.close(fig)
        output.seek(0)

        return output


class ExcelTemplateImageGenerator:
    """Excelテンプレート方式の画像生成クラス"""

    def __init__(self, template_path: str = None):
        """
        Args:
            template_path: Excelテンプレートパス
        """
        self.template_path = template_path or EXCEL_TEMPLATE_PATH

    def generate_from_template(
        self,
        data: List[Dict],
        sheet_name: str = "総合1つ",
        ranking_name: str = "",
        year: int = None,
        sample_size: int = None
    ) -> Optional[BytesIO]:
        """テンプレートにデータを埋め込んでExcelを生成

        Args:
            data: ランキングデータ
            sheet_name: 使用するシート名
            ranking_name: ランキング名
            year: 年度
            sample_size: 回答者数

        Returns:
            BytesIO（Excelファイル）
        """
        if not os.path.exists(self.template_path):
            logger.error(f"テンプレートが見つかりません: {self.template_path}")
            return None

        try:
            wb = openpyxl.load_workbook(self.template_path)

            if sheet_name not in wb.sheetnames:
                logger.error(f"シートが見つかりません: {sheet_name}")
                return None

            ws = wb[sheet_name]

            # データ埋め込み（シート構造に依存）
            # 「総合1つ」シートの場合:
            # Row3: タイトル
            # Row4: ランキング名
            # Row5: ヘッダー
            # Row6-10: データ

            if sheet_name == "総合1つ":
                # タイトル（Row3, Col D）
                if year:
                    ws['D3'] = f"　　　　　{year}年 オリコン顧客満足度®調査 総合ランキング"

                # ランキング名（Row4, Col C）
                subtitle = ranking_name
                if sample_size:
                    subtitle += f" (回答者数：{sample_size:,}名)"
                ws['C4'] = subtitle

                # データ（Row6-10, Col C-E）
                for i, entry in enumerate(data[:5]):
                    row = 6 + i
                    ws[f'C{row}'] = f"{entry.get('rank', '')}位"
                    ws[f'D{row}'] = entry.get('company', '')
                    score = entry.get('score')
                    ws[f'E{row}'] = score if score is not None else ''

            # BytesIOに保存
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            return output

        except Exception as e:
            logger.error(f"Excel生成エラー: {e}")
            return None


# ========================================
# 便利関数
# ========================================
def generate_ranking_image(
    ranking_name: str,
    year: int,
    data: List[Dict],
    image_type: str = "overall",
    prev_data: List[Dict] = None,
    **kwargs
) -> Optional[BytesIO]:
    """ランキング表画像を生成（簡易インターフェース）

    Args:
        ranking_name: ランキング名
        year: 年度
        data: ランキングデータ
        image_type: "overall", "comparison", "multi"
        prev_data: 前年データ（comparison時）
        **kwargs: その他オプション

    Returns:
        BytesIO（PNG画像）
    """
    generator = TableImageGenerator(ranking_name, year)

    if image_type == "overall":
        return generator.generate_overall_table_image(data, **kwargs)
    elif image_type == "comparison":
        return generator.generate_comparison_table_image(data, prev_data, **kwargs)
    elif image_type == "multi":
        return generator.generate_multi_table_image(data, **kwargs)

    return None


# ========================================
# デバッグ用
# ========================================
if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8')

    print("=== Image Generator Test ===")

    # テストデータ
    test_data = [
        {"rank": 1, "company": "SBI証券", "score": 68.9},
        {"rank": 2, "company": "楽天証券", "score": 68.0},
        {"rank": 3, "company": "マネックス証券", "score": 67.5},
        {"rank": 4, "company": "松井証券", "score": 66.0},
        {"rank": 5, "company": "auカブコム証券", "score": 65.5},
    ]

    prev_data = [
        {"rank": 1, "company": "SBI証券", "score": 68.5},
        {"rank": 2, "company": "マネックス証券", "score": 67.0},
        {"rank": 3, "company": "楽天証券", "score": 66.5},
        {"rank": 4, "company": "松井証券", "score": 65.5},
    ]

    # 総合ランキング画像
    generator = TableImageGenerator("ネット証券", 2026)

    img1 = generator.generate_overall_table_image(test_data, sample_size=5000)
    if img1:
        with open("test_overall.png", "wb") as f:
            f.write(img1.getvalue())
        print("[OK] test_overall.png generated")

    # 前年比較画像
    img2 = generator.generate_comparison_table_image(test_data, prev_data)
    if img2:
        with open("test_comparison.png", "wb") as f:
            f.write(img2.getvalue())
        print("[OK] test_comparison.png generated")

    # 複数表画像
    multi_data = {
        "取引手数料": test_data[:3],
        "取引ツール": test_data[:3],
        "情報提供": test_data[:3],
    }
    img3 = generator.generate_multi_table_image(multi_data)
    if img3:
        with open("test_multi.png", "wb") as f:
            f.write(img3.getvalue())
        print("[OK] test_multi.png generated")

    print("\n=== All Tests Completed ===")
